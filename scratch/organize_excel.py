import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

file_path = 'app_strings_translations.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# 1. Sort Data Alphabetically by Key
data = []
header = [cell.value for cell in ws[1]]
for row in ws.iter_rows(min_row=2, values_only=True):
    data.append(row)

# The first column is 'key'
data.sort(key=lambda x: str(x[0]) if x[0] else "")

# Write sorted data back
for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 2. Format Header Row
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

for col_idx in range(1, len(header) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# 3. Column Auto-Sizing
for col_idx in range(1, len(header) + 1):
    col_letter = get_column_letter(col_idx)
    max_length = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
        for cell in row:
            if cell.value:
                # Add a little padding to the length
                max_length = max(max_length, len(str(cell.value)))
    
    # Cap width to avoid extremely wide columns
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[col_letter].width = adjusted_width

# Set row alignment for readability (Wrap text for long strings)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = wrap_alignment

# 4. Freeze Panes
ws.freeze_panes = "A2" # Freezes row 1

# 5. Auto-Filters
last_col_letter = get_column_letter(len(header))
ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

wb.save(file_path)
print("Excel file successfully organized!")
