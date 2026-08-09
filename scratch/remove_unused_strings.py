import openpyxl, os, re

# 1. Identify all keys
wb = openpyxl.load_workbook('app_strings_translations.xlsx')
ws = wb.active
header = [c.value for c in ws[1]]
key_idx = header.index('key')

# Map key to its row index (1-based)
all_keys_map = {}
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    key = row[key_idx]
    if key:
        all_keys_map[key] = i

used_keys = set()
for root, dirs, files in os.walk('app/src/main'):
    for file in files:
        if file == 'strings.xml':
            continue
        if file.endswith('.xml') or file.endswith('.kt') or file.endswith('.java'):
            filepath = os.path.join(root, file)
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
                remaining = set(all_keys_map.keys()) - used_keys
                for key in remaining:
                    # check for @string/key, @+string/key, R.string.key, "key"
                    if f'string/{key}' in content or f'R.string.{key}' in content or f'"{key}"' in content:
                        used_keys.add(key)

unused_keys = set(all_keys_map.keys()) - used_keys
print(f'Found {len(unused_keys)} unused keys.')

# We must delete rows from bottom to top to avoid shifting row indices
rows_to_delete = sorted([all_keys_map[k] for k in unused_keys], reverse=True)

for r in rows_to_delete:
    key_name = ws.cell(row=r, column=key_idx+1).value
    print(f"Deleting row {r} (Key: {key_name})")
    ws.delete_rows(r)

wb.save('app_strings_translations.xlsx')
print(f"Successfully deleted {len(rows_to_delete)} rows from Excel.")
