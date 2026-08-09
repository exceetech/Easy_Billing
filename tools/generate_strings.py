#!/usr/bin/env python3
"""
Regenerates every res/values*/strings.xml file in the app from a single
spreadsheet: app_strings_translations.xlsx (in the project root, next to
this tools/ folder).

USAGE (whenever you add/change/translate a string):
    1. Open app_strings_translations.xlsx.
    2. To add a new string: add a new row with a unique `key` (snake_case,
       no spaces), a `section` (just a label for grouping — e.g. "Invoice",
       "Settings"), and the `english` text. Fill in whichever language
       columns you already have translations for; leave the rest blank.
    3. To change existing text: edit the cell directly. Every language
       whose cell you leave blank simply falls back to the English text
       at runtime (this is normal Android behavior — you do NOT need to
       fill in every language for every row).
    4. Save the spreadsheet, then run:
           python3 tools/generate_strings.py
       from the project root (Easy_Billing/). It rewrites all 6
       strings.xml files. Review the diff, then build normally.

No manual XML editing required — the spreadsheet is the single source of
truth. Do not hand-edit the strings.xml files directly; regenerate them
from the spreadsheet instead, or your edit will be overwritten next run.

Columns expected in the active sheet:
    key | section | english | hindi | kannada | malayalam | tamil | telugu | status | source_locations

"status" and "source_locations" are informational only (not written to XML).
"""
import re
import sys
import collections
from pathlib import Path
from xml.sax.saxutils import escape

# Matches a valid Java/Android format specifier starting at a '%':
# %s, %d, %.2f, %1$s, %2$d, %%, etc. Anything else that starts with '%'
# is NOT a real specifier and must be escaped as a literal '%%', or
# Android's resource compiler rejects the whole file at build time with
# an opaque "Can not extract resource" error (aaptcompiler validates
# every string for format-specifier syntax at compile time, regardless
# of whether the string is ever actually used as a format string).
_FORMAT_SPEC_RE = re.compile(r"%(?:[0-9]+\$)?[-+ 0,(#]*[0-9]*(?:\.[0-9]+)?[a-zA-Z%]")


def escape_percent(text: str) -> str:
    out = []
    i, n = 0, len(text)
    while i < n:
        if text[i] == "%":
            m = _FORMAT_SPEC_RE.match(text, i)
            if m:
                out.append(m.group(0))
                i = m.end()
                continue
            out.append("%%")
            i += 1
        else:
            out.append(text[i])
            i += 1
    return "".join(out)

try:
    import openpyxl
except ImportError:
    print("Missing dependency. Install it with:\n    pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "app_strings_translations.xlsx"
RES_DIR = ROOT / "app" / "src" / "main" / "res"

# Column key -> (values-<qualifier> dir name or None for default)
LANG_COLUMNS = {
    "english": None,       # values/strings.xml
    "hindi": "hi",
    "kannada": "kn",
    "malayalam": "ml",
    "tamil": "ta",
    "telugu": "te",
}


def xml_escape_value(text: str) -> str:
    """Escape a string for use as Android string resource content.

    Handles XML entities plus Android's own special characters: apostrophes
    and double quotes must be escaped even though XML itself doesn't
    require it, and newlines become the \\n escape Android expects.
    """
    if text is None:
        text = ""
    text = str(text)
    text = escape_percent(text)  # lone % -> %% (must run before & escaping, harmless either order here)
    text = escape(text)  # & < >
    text = text.replace("'", "\\'").replace('"', '\\"')
    text = text.replace("\r\n", "\\n").replace("\n", "\\n")
    # A string starting with '@' or '?' is parsed by Android's resource
    # compiler as a resource/theme-attribute REFERENCE (e.g. "@string/foo"),
    # not literal text — even if what follows isn't valid reference syntax,
    # which can desync the resource-table parser badly enough to report
    # unrelated "Can not extract resource" failures elsewhere in the same
    # file. A leading backslash forces it to be treated as literal text.
    if text.startswith("@") or text.startswith("?"):
        text = "\\" + text
        
    # Preserve leading and trailing spaces by wrapping in double quotes
    if text.startswith(" ") or text.endswith(" "):
        text = f'"{text}"'
        
    return text


def load_rows():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: i for i, name in enumerate(header)}

    required = ["key", "section", "english"]
    for r in required:
        if r not in col_idx:
            print(f"ERROR: spreadsheet is missing required column '{r}'", file=sys.stderr)
            sys.exit(1)

    rows = []
    seen_keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        key = row[col_idx["key"]]
        if not key or not str(key).strip():
            continue
        key = str(key).strip()
        if key in seen_keys:
            print(f"WARNING: duplicate key '{key}' — later row overrides earlier one", file=sys.stderr)
        seen_keys.add(key)
        section = (row[col_idx["section"]] or "General")
        entry = {"key": key, "section": str(section).strip() or "General"}
        for lang_col in LANG_COLUMNS:
            idx = col_idx.get(lang_col)
            val = row[idx] if idx is not None and idx < len(row) else None
            # Only check truthiness of stripped string to avoid saving pure-whitespace keys
            if val is not None and str(val).strip():
                entry[lang_col] = str(val).strip()
            else:
                entry[lang_col] = None
        rows.append(entry)
    return rows


def write_strings_xml(path: Path, rows, lang_col: str, include_all_keys_for_default: bool):
    """Write one strings.xml. For the default (English) file every row is
    written. For a translated-language file, only rows with a non-empty
    value in that language are written — Android automatically falls back
    to the default value for any key missing from a values-<qualifier>
    file, so omitting untranslated rows is correct, not a bug.
    """
    by_section = collections.OrderedDict()
    for r in rows:
        text = r["english"] if include_all_keys_for_default else r.get(lang_col)
        if text is None:
            continue
        by_section.setdefault(r["section"], []).append((r["key"], text))

    lines = ['<?xml version="1.0" encoding="utf-8"?>', "<resources>", ""]
    for section, items in by_section.items():
        lines.append(f"    <!-- {section} -->")
        for key, text in items:
            lines.append(f'    <string name="{key}">{xml_escape_value(text)}</string>')
        lines.append("")
    lines.append("</resources>")
    lines.append("")

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {path.relative_to(ROOT)}  ({sum(len(v) for v in by_section.values())} strings)")


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found.", file=sys.stderr)
        sys.exit(1)

    rows = load_rows()
    print(f"Loaded {len(rows)} string rows from {XLSX_PATH.name}\n")

    for lang_col, qualifier in LANG_COLUMNS.items():
        is_default = qualifier is None
        out_dir = RES_DIR / "values" if is_default else RES_DIR / f"values-{qualifier}"
        write_strings_xml(out_dir / "strings.xml", rows, lang_col, include_all_keys_for_default=is_default)

    print("\nDone. Review the diff (git diff app/src/main/res/values*/strings.xml) before committing.")


if __name__ == "__main__":
    main()
